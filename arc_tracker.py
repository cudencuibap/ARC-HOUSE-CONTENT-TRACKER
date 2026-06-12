#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Arc House Content Tracker
-------------------------
Theo doi noi dung video/bai viet tren community.arc.io de toi uu diem contribution.

Cach hoat dong (da reverse-engineer tu site that - xem CLAUDE.md):
- Site la Next.js + GraphQL (nen tang Gradual). Endpoint: POST /api/graphql
- Moi request can header chong bot `x-nonce` (HMAC + timestamp) sinh boi JS that
  cua site -> tool goi `nonce_gen.js` (Node) de tao nonce tuoi cho tung request.
- Cookie dang nhap doc tu /config/cookie.txt (KHONG hard-code).
- User-Agent PHAI giong luc lay cookie (cf_clearance gan voi UA) -> dung UA mobile.

Lay du lieu:
- Content public: query getContentPageContents + getPaginatedCollectionPageContents
  (trang chinh + cac collection) -> luu Excel.
- Lich su da xem: query myContributionLogs -> cac log "Watch a Video"/"Read Content"
  co title + ngay that -> tu danh dau "Da xem".
- Lifetime points: query myTenantUser.totalContributionPoints.

Chay trong Docker. Data luu o /data (mount ra host).
"""

import os
import re
import sys
import json
import uuid
import subprocess
import datetime as dt

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ----------------------------- Cau hinh -----------------------------
BASE_URL  = "https://community.arc.io"
GQL_URL   = f"{BASE_URL}/api/graphql"
TENANT_ID = "688b1c42d1f8bff20fcac7a4"   # tenant "Arc House" (public, lay tu __NEXT_DATA__)

# x-client: phải khớp với chuỗi dùng để sinh nonce (server kiểm HMAC theo x-client).
# Giá trị này chỉ cần nhất quán giữa nonce và header; lấy từ build hiện tại của site.
X_CLIENT = os.environ.get(
    "ARC_X_CLIENT",
    "Gradual - v3.99.0-prod-563002e-5330250a3d7d628bf4a80828a6213a8f")

# User-Agent mobile (giống lúc lấy cookie / cf_clearance). Sai UA -> Cloudflare/403.
USER_AGENT = os.environ.get(
    "ARC_USER_AGENT",
    "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/148.0.0.0 Mobile Safari/537.36")

# Cac nguon content can cao (trang chinh + collections)
COLLECTION_SLUGS = ["architects", "stablecoin-101-2025-10-25"]

DATA_DIR    = os.environ.get("ARC_DATA_DIR", "/data")
CONFIG_DIR  = os.environ.get("ARC_CONFIG_DIR", "/config")
APP_DIR     = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(DATA_DIR, "arc-content.xlsx")
DEBUG_DIR   = os.path.join(DATA_DIR, "debug")
LANG_PATH   = os.path.join(CONFIG_DIR, "language.txt")   # "en" hoac "vi"

# nonce: script Node + file chunk JS (co the bi sua/refresh khi site deploy lai)
NONCE_SCRIPT = os.path.join(APP_DIR, "nonce_gen.js")
CHUNK_VENDOR = os.path.join(APP_DIR, "nonce_chunk.js")        # ban di kem image
CHUNK_DATA   = os.path.join(DATA_DIR, "nonce_chunk.js")       # ban refresh (ghi duoc)

# Quota theo rule Arc
VIDEO_DAILY_MAX   = 4   # 4 diem/video, toi da 4 video/24h
ARTICLE_DAILY_MAX = 5   # 2 diem/bai,   toi da 5 bai/24h
VIDEO_POINTS      = 4
ARTICLE_POINTS    = 2

COLUMNS = ["ID", "Type", "Title", "URL", "Points",
           "Status", "DateAdded", "DateViewed", "Note", "Blacklist"]

# ----------------------------- Tien ich -----------------------------
class C:
    R = "\033[91m"; G = "\033[92m"; Y = "\033[93m"; B = "\033[94m"
    M = "\033[95m"; CY = "\033[96m"; GR = "\033[90m"; W = "\033[0m"; BOLD = "\033[1m"

# ----------------------------- i18n (EN/VI) -----------------------------
# QUAN TRONG: cac string o day chi la DISPLAY value (dich theo ngon ngu).
# Gia tri INTERNAL trong Excel ("Chua xem"/"Da xem"/"Yes"/"Video"/"Article")
# KHONG nam o day -> giu nguyen, khong bao gio doi -> data cu tuong thich 100%.
TRANSLATIONS = {
    "en": {
        # startup
        "startup":            "Starting Arc House Tracker...",
        "excel_path":         "Excel: {path}",
        "error_nonce":        "Could not generate x-nonce (need Node + nonce_gen.js). Stopping.",
        "cookie_loaded":      "Cookie + nonce OK -> syncing history + points.",
        "no_cookie_public":   "No cookie -> scraping public content only.",
        # sync
        "sync_fetching":      "Loading content from Arc House (main page + collections)...",
        "sync_found":         "Found {n} content items.",
        "sync_collection_err":"  Collection {slug} error: {e}",
        "sync_logs_read":     "Read {logs} logs -> {viewed} content viewed/read.",
        "sync_marked_viewed": "Marked {n} items as viewed/read from real history.",
        "sync_new_items":     "Added {n} new items.",
        "error_no_cookie":    "No cookie -> skipping history sync (see README).",
        "error_graphql":      "Could not read history (GraphQL): {e}",
        "error_points":       "Could not fetch lifetime points: {e}",
        "error_network":      "Content scrape error: {e}",
        "error_excel":        "Could not read old Excel ({e}), creating new.",
        # dashboard
        "dashboard_title":            "ARC HOUSE CONTENT TRACKER",
        "dashboard_lifetime_section": "LIFETIME",
        "dashboard_lifetime_points":  "Lifetime points",
        "dashboard_videos_progress":  "Videos",
        "dashboard_articles_progress":"Articles",
        "dashboard_viewed_suffix":    "viewed",
        "dashboard_read_suffix":      "read",
        "dashboard_today_section":    "TODAY ({date})",
        "dashboard_video_quota_today":   "Videos",
        "dashboard_article_quota_today": "Articles",
        "dashboard_daily_active":     "Daily Active",
        "dashboard_points_today":     "Points today (from real log): {pts}",
        "dashboard_max_potential":    "STILL EARNABLE",
        "dashboard_earnable":         "{vslot} video x4 + {aslot} article x2 = {earnable} pts",
        # recommend
        "recommend_header":        "TODAY'S RECOMMENDATIONS",
        "recommend_maxed_out":     "Daily view/read points MAXED out! Come back tomorrow.",
        "recommend_videos_left":   "[Video] {vslot} slot(s) left (4pts each):",
        "recommend_no_videos":     "  (No unwatched videos left)",
        "recommend_articles_left": "[Article] {aslot} slot(s) left (2pts each):",
        "recommend_no_articles":   "  (No unread articles left)",
        "recommend_earnable":      "Still earnable: {pts} pts today",
        # report
        "report_header":       "REPORT DEAD LINKS",
        "report_no_items":     "  (No recommended items to report)",
        "report_select_prompt":"Enter dead-link item numbers (e.g. 1,3,5), Enter to skip: ",
        "report_nothing":      "Nothing reported.",
        "report_confirmed":    "Blacklisted {n} item(s) (saved to Excel, won't be recommended again).",
        "report_none_new":     "No new items were blacklisted.",
        # menu
        "menu_title":     "MENU",
        "menu_recommend": "Today's recommendations (refresh + sync)",
        "menu_report":    "Report dead links (blacklist)",
        "menu_dashboard": "Dashboard",
        "menu_exit":      "Exit",
        "prompt_choose":  "Choose: ",
        "prompt_invalid": "Invalid choice.",
        "exit_message":   "Bye!",
        # status (display only)
        "status_not_viewed": "Not viewed",
        "status_viewed":     "Viewed",
        "status_blacklisted":"Blacklisted",
        # language picker
        "lang_prompt": "Select language / Chon ngon ngu:\n  1. English\n  2. Tieng Viet",
        "lang_choose": "Choose (1/2): ",
        "lang_set":    "Language set to English.",
    },
    "vi": {
        # startup
        "startup":            "Khoi dong Arc House Tracker...",
        "excel_path":         "Excel: {path}",
        "error_nonce":        "Khong sinh duoc x-nonce (can Node + nonce_gen.js). Dung lai.",
        "cookie_loaded":      "Da nap cookie + nonce OK -> dong bo lich su + points.",
        "no_cookie_public":   "Chua co cookie -> chi cao content public.",
        # sync
        "sync_fetching":      "Dang tai content tu Arc House (trang chinh + collections)...",
        "sync_found":         "Tim thay {n} item content.",
        "sync_collection_err":"  Collection {slug} loi: {e}",
        "sync_logs_read":     "Doc {logs} log -> {viewed} content da xem/doc.",
        "sync_marked_viewed": "Danh dau {n} item da xem/doc tu lich su that.",
        "sync_new_items":     "Them {n} item moi.",
        "error_no_cookie":    "Chua co cookie -> bo qua dong bo lich su (xem README).",
        "error_graphql":      "Khong doc duoc lich su (GraphQL): {e}",
        "error_points":       "Khong lay duoc lifetime points: {e}",
        "error_network":      "Loi cao content: {e}",
        "error_excel":        "Khong doc duoc Excel cu ({e}), tao moi.",
        # dashboard
        "dashboard_title":            "ARC HOUSE CONTENT TRACKER",
        "dashboard_lifetime_section": "LIFETIME",
        "dashboard_lifetime_points":  "Lifetime points",
        "dashboard_videos_progress":  "Videos",
        "dashboard_articles_progress":"Articles",
        "dashboard_viewed_suffix":    "da xem",
        "dashboard_read_suffix":      "da doc",
        "dashboard_today_section":    "HOM NAY ({date})",
        "dashboard_video_quota_today":   "Videos",
        "dashboard_article_quota_today": "Articles",
        "dashboard_daily_active":     "Daily Active",
        "dashboard_points_today":     "Diem hom nay (tu log that): {pts}",
        "dashboard_max_potential":    "CON CO THE KIEM",
        "dashboard_earnable":         "{vslot} video x4 + {aslot} article x2 = {earnable} diem",
        # recommend
        "recommend_header":        "DE XUAT HOM NAY",
        "recommend_maxed_out":     "Da MAX diem xem/doc hom nay! Quay lai mai.",
        "recommend_videos_left":   "[Video] Con {vslot} slot (4d/cai):",
        "recommend_no_videos":     "  (Het video chua xem)",
        "recommend_articles_left": "[Article] Con {aslot} slot (2d/cai):",
        "recommend_no_articles":   "  (Het bai chua doc)",
        "recommend_earnable":      "Con co the kiem: {pts} diem hom nay",
        # report
        "report_header":       "REPORT LINK DIE",
        "report_no_items":     "  (Khong co item nao dang de xuat de report)",
        "report_select_prompt":"Nhap so item link die (vd 1,3,5), Enter de bo qua: ",
        "report_nothing":      "Khong report gi.",
        "report_confirmed":    "Da blacklist {n} item (da luu Excel, se khong de xuat lai).",
        "report_none_new":     "Khong co item moi nao duoc blacklist.",
        # menu
        "menu_title":     "MENU",
        "menu_recommend": "De xuat hom nay (refresh + dong bo)",
        "menu_report":    "Report link die (blacklist)",
        "menu_dashboard": "Dashboard",
        "menu_exit":      "Thoat",
        "prompt_choose":  "Chon: ",
        "prompt_invalid": "Lua chon khong hop le.",
        "exit_message":   "Bye!",
        # status (display only)
        "status_not_viewed": "Chua xem",
        "status_viewed":     "Da xem",
        "status_blacklisted":"Da blacklist",
        # language picker
        "lang_prompt": "Select language / Chon ngon ngu:\n  1. English\n  2. Tieng Viet",
        "lang_choose": "Choose / Chon (1/2): ",
        "lang_set":    "Da chon Tieng Viet.",
    },
}

LANG = "en"   # set boi load_language()/select_language() khi khoi dong

def t(key, **kwargs):
    """Tra string theo LANG hien tai; fallback 'en' neu key thieu."""
    s = TRANSLATIONS.get(LANG, {}).get(key)
    if s is None:
        s = TRANSLATIONS["en"].get(key, key)
    return s.format(**kwargs) if kwargs else s

def load_language():
    """Doc ngon ngu da luu tu config/language.txt. Tra ve 'en'/'vi' hoac None."""
    try:
        if os.path.exists(LANG_PATH):
            with open(LANG_PATH, "r", encoding="utf-8") as f:
                v = f.read().strip().lower()
            if v in TRANSLATIONS:
                return v
    except Exception:
        pass
    return None

def select_language():
    """Set bien global LANG. Lan dau (chua co file) thi hoi user va luu lai.
    Non-TTY (CI/pipe) ma chua co file -> mac dinh 'en', khong hoi."""
    global LANG
    saved = load_language()
    if saved:
        LANG = saved
        return
    if not sys.stdin.isatty():
        LANG = "en"
        return
    print(TRANSLATIONS["en"]["lang_prompt"])
    while True:
        choice = input(TRANSLATIONS["en"]["lang_choose"]).strip()
        if choice == "1":
            LANG = "en"; break
        if choice == "2":
            LANG = "vi"; break
        print("Invalid / Khong hop le.")
    try:
        os.makedirs(CONFIG_DIR, exist_ok=True)
        with open(LANG_PATH, "w", encoding="utf-8") as f:
            f.write(LANG)
    except Exception:
        pass
    print(f"{C.G}{t('lang_set')}{C.W}")

def status_display(d):
    """Status INTERNAL -> DISPLAY (translated). Khong doi gia tri trong Excel."""
    return t("status_viewed") if d.get("Status") == "Da xem" else t("status_not_viewed")

def today_str():
    return dt.date.today().isoformat()

def norm(s):
    """Chuan hoa text de so khop title."""
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def load_cookie():
    """Doc cookie tu env hoac file. Tra ve chuoi cookie hoac None."""
    env = os.environ.get("ARC_COOKIE", "").strip()
    if env:
        return env
    path = os.path.join(CONFIG_DIR, "cookie.txt")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            val = f.read().strip()
        if val and not val.startswith("#"):
            return val
    return None

def make_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "accept": "*/*",
        "accept-language": "en-US,en;q=0.9",
        "content-type": "application/json",
        "origin": BASE_URL,
        "referer": f"{BASE_URL}/home/contributors/my-contributions",
        "x-client": X_CLIENT,
        "x-locale": "en",
        "x-tenant-id": TENANT_ID,
    })
    cookie = load_cookie()
    if cookie:
        s.headers["Cookie"] = cookie
    return s, bool(cookie)

# ----------------------------- Nonce chong bot -----------------------------
def _chunk_path():
    return CHUNK_DATA if os.path.exists(CHUNK_DATA) else CHUNK_VENDOR

def gen_nonce(chunk_path=None):
    """Goi Node de sinh x-nonce tuoi tu chinh JS cua site."""
    chunk_path = chunk_path or _chunk_path()
    env = dict(os.environ, ARC_NONCE_CHUNK=chunk_path)
    out = subprocess.check_output(
        ["node", NONCE_SCRIPT, X_CLIENT], env=env, text=True,
        stderr=subprocess.DEVNULL, timeout=30)
    return out.strip()

def refresh_nonce_chunk(session):
    """Tai lai chunk JS chua generateNonce tu site (khi site deploy lai)."""
    try:
        r = session.get(f"{BASE_URL}/home/contributors/my-contributions", timeout=30)
        html = r.text
        srcs = re.findall(r'<script[^>]+src="([^"]+\.js)"', html)
        for u in srcs:
            full = u if u.startswith("http") else BASE_URL + u
            try:
                js = session.get(full, timeout=30).text
            except Exception:
                continue
            if "generateNonce" in js and "36035:function" in js:
                os.makedirs(DATA_DIR, exist_ok=True)
                with open(CHUNK_DATA, "w", encoding="utf-8") as f:
                    f.write(js)
                return True
    except Exception as e:
        print(f"{C.R}Khong refresh duoc nonce chunk: {e}{C.W}")
    return False

def ensure_nonce(session):
    """Dam bao sinh duoc nonce; neu khong, thu refresh chunk tu site."""
    try:
        if gen_nonce():
            return True
    except Exception:
        pass
    print(f"{C.Y}Nonce loi -> dang lam moi tu site...{C.W}")
    if refresh_nonce_chunk(session):
        try:
            return bool(gen_nonce(CHUNK_DATA))
        except Exception as e:
            print(f"{C.R}Van loi sau khi refresh: {e}{C.W}")
    return False

# ----------------------------- GraphQL -----------------------------
def gql(session, operation, query, variables):
    """POST 1 query GraphQL voi nonce tuoi. Tra ve dict 'data' hoac raise."""
    headers = {
        "x-request-id": str(uuid.uuid4()),
        "x-session-id": str(uuid.uuid4()),
        "x-nonce": gen_nonce(),
    }
    payload = {"operationName": operation, "query": query, "variables": variables}
    r = session.post(GQL_URL, data=json.dumps(payload), headers=headers, timeout=40)
    try:
        body = r.json()
    except Exception:
        raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
    if body.get("errors"):
        raise RuntimeError(body["errors"][0].get("message", "GraphQL error"))
    return body.get("data") or {}

# Chi lay field can thiet (id/slug/title/typename) cho gon
_CONTENT_REC = ("... on Blog { id slug title __typename } "
                "... on LVTenantVideo { id slug title __typename }")

Q_CONTENT_PAGE = (
    "query GetContentPageContents($tenantId: String!, $skip: Int, $limit: Int) {"
    "  contents: getContentPageContents(tenantId: $tenantId, skip: $skip, limit: $limit) {"
    "    totalCount records { %s } } }" % _CONTENT_REC)

Q_COLLECTION = (
    "query GetPaginatedCollectionPageTenantContents($tenantId: String!, $collectionSlug: String!, $skip: Int, $limit: Int) {"
    "  contents: getPaginatedCollectionPageContents(tenantId: $tenantId, collectionSlug: $collectionSlug, skip: $skip, limit: $limit) {"
    "    totalCount records { %s } } }" % _CONTENT_REC)

Q_POINTS = (
    "query MyContributionInfo($tenantId: MongoID!) {"
    "  tenantUser: myTenantUser(tenantId: $tenantId) {"
    "    id totalContributionPoints totalContributionBadgeCount } }")

Q_LOGS = (
    "query MyContributionLogs($tenantId: ID!, $skip: Int, $limit: Int) {"
    "  contributionLogs: myContributionLogs(tenantId: $tenantId, skip: $skip, limit: $limit) {"
    "    id point count roleType logType occurredAt notes"
    "    contributionRole { id name }"
    "    group { ... on Blog { id title } ... on LVTenantVideo { id title }"
    "            ... on ForumPost { id title } ... on Meetup { id name } ... on Club { id name } } } }")

PAGE = 50  # limit toi da server chap nhan

def _record_to_item(rec):
    tp = rec.get("__typename", "")
    slug = rec.get("slug", "")
    title = (rec.get("title") or slug.replace("-", " ").title()).strip()
    if tp == "LVTenantVideo":
        ctype = "Video"; url = f"{BASE_URL}/public/videos/{slug}"
    else:
        ctype = "Article"; url = f"{BASE_URL}/public/blogs/{slug}"
    rid = f"{ctype.lower()}:{slug}"
    return {
        "ID": rid, "Type": ctype, "Title": title, "URL": url,
        "Points": VIDEO_POINTS if ctype == "Video" else ARTICLE_POINTS,
        "Status": "Chua xem", "DateAdded": today_str(),
        "DateViewed": "", "Note": "", "Blacklist": "",
    }

# ----------------------------- Cao content public -----------------------------
def fetch_public_content(session):
    print(f"{C.CY}{t('sync_fetching')}{C.W}")
    items = {}

    def add(recs):
        for rec in recs:
            if not rec.get("slug"):
                continue
            it = _record_to_item(rec)
            items.setdefault(it["ID"], it)

    # 1) Trang content chinh (co pagination)
    skip = 0
    while True:
        d = gql(session, "GetContentPageContents", Q_CONTENT_PAGE,
                {"tenantId": TENANT_ID, "skip": skip, "limit": PAGE}).get("contents")
        if not d:
            break
        recs = d["records"]; add(recs)
        skip += len(recs)
        if not recs or skip >= d["totalCount"]:
            break

    # 2) Cac collection
    for slug in COLLECTION_SLUGS:
        s2 = 0
        while True:
            try:
                d = gql(session, "GetPaginatedCollectionPageTenantContents", Q_COLLECTION,
                        {"tenantId": TENANT_ID, "collectionSlug": slug, "skip": s2, "limit": PAGE}).get("contents")
            except Exception as e:
                print(f"{C.Y}{t('sync_collection_err', slug=slug, e=e)}{C.W}")
                break
            if not d:
                break
            recs = d["records"]; add(recs)
            s2 += len(recs)
            if not recs or s2 >= d["totalCount"]:
                break

    print(f"{C.G}{t('sync_found', n=len(items))}{C.W}")
    return list(items.values())

# ----------------------------- Lich su da xem + points -----------------------------
def fetch_contribution_logs(session):
    """Lay TOAN BO contribution logs (co pagination)."""
    logs = []; skip = 0
    while True:
        arr = gql(session, "MyContributionLogs", Q_LOGS,
                  {"tenantId": TENANT_ID, "skip": skip, "limit": PAGE}).get("contributionLogs")
        if not arr:
            break
        logs += arr; skip += len(arr)
        if len(arr) < PAGE:
            break
    return logs

def fetch_lifetime_points(session):
    try:
        tu = gql(session, "MyContributionInfo", Q_POINTS, {"tenantId": TENANT_ID}).get("tenantUser") or {}
        return tu.get("totalContributionPoints")
    except Exception as e:
        print(f"{C.Y}{t('error_points', e=e)}{C.W}")
        return None

def fetch_viewed_titles(session, has_cookie):
    """
    Tra ve dict { title_da_chuan_hoa: ngay_xem } cho cac content user DA xem/doc,
    lay tu myContributionLogs (role 'Watch a Video' / 'Read Content').
    Khong co cookie / loi -> tra ve {}.
    """
    if not has_cookie:
        print(f"{C.Y}{t('error_no_cookie')}{C.W}")
        return {}
    os.makedirs(DEBUG_DIR, exist_ok=True)
    try:
        logs = fetch_contribution_logs(session)
    except Exception as e:
        print(f"{C.R}{t('error_graphql', e=e)}{C.W}")
        return {}

    with open(os.path.join(DEBUG_DIR, "contribution_logs.json"), "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)

    viewed = {}
    for l in logs:
        role = (l.get("contributionRole") or {}).get("name") or ""
        if role not in ("Watch a Video", "Read Content"):
            continue
        g = l.get("group") or {}
        title = g.get("title") or g.get("name")
        if not title:
            continue
        date = (l.get("occurredAt") or "")[:10]
        k = norm(title)
        if k and (k not in viewed or date > viewed[k]):
            viewed[k] = date
    print(f"{C.G}{t('sync_logs_read', logs=len(logs), viewed=len(viewed))}{C.W}")
    return viewed

def apply_viewed(data, viewed):
    """Danh dau Da xem + ngay that. KHONG bao gio reset 'Da xem' khi refresh."""
    vitems = list(viewed.items())
    matched = 0
    for d in data:
        k = norm(d["Title"])
        date = viewed.get(k)
        if date is None:
            for vk, vd in vitems:           # fuzzy: title log la tien to/hau to cua title content
                if len(vk) >= 12 and (vk.startswith(k) or k.startswith(vk)):
                    date = vd; break
        if date:
            d["Status"] = "Da xem"
            d["DateViewed"] = date
            matched += 1
        # neu khong match nhung truoc do da "Da xem" -> giu nguyen (khong downgrade)
    return matched

# ----------------------------- Excel I/O -----------------------------
def load_excel():
    if not os.path.exists(EXCEL_PATH):
        return []
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Content"]
        rows = []
        headers = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            d = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
            for col in COLUMNS:
                d.setdefault(col, "")
                if d[col] is None:
                    d[col] = ""
            rows.append(d)
        return rows
    except Exception as e:
        print(f"{C.Y}{t('error_excel', e=e)}{C.W}")
        return []

def save_excel(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Content"

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    done_fill  = PatternFill("solid", fgColor="C6EFCE")   # xanh la nhat (Da xem)
    black_fill = PatternFill("solid", fgColor="D9D9D9")   # xam nhat (Blacklist)

    ws.append(COLUMNS)
    for i, _ in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=i)
        c.fill = head_fill; c.font = head_font
        c.alignment = Alignment(vertical="center")

    # sap xep: chua xem len truoc, moi nhat truoc
    data_sorted = sorted(
        data,
        key=lambda d: (d.get("Status") == "Da xem", d.get("Type", ""), d.get("Title", "")),
    )
    ncol = len(COLUMNS)
    for d in data_sorted:
        ws.append([d.get(c, "") for c in COLUMNS])
        row = ws.max_row
        # Uu tien xanh (Da xem) > xam (Blacklist) > trang (Chua xem chua report)
        fill = None
        if d.get("Status") == "Da xem":
            fill = done_fill                   # TO XANH CA DONG A:J
        elif str(d.get("Blacklist", "")).strip().lower() == "yes":
            fill = black_fill                  # TO XAM CA DONG A:J (da report link die)
        if fill is not None:
            for col in range(1, ncol + 1):
                ws.cell(row=row, column=col).fill = fill

    widths = [30, 9, 52, 48, 7, 10, 12, 12, 16, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)

# ----------------------------- Sync / merge -----------------------------
def sync(session, has_cookie):
    """Cao content + dong bo lich su. Tra ve (data, logs, points)."""
    existing = load_excel()
    by_id = {d["ID"]: d for d in existing}

    try:
        fresh = fetch_public_content(session)
    except Exception as e:
        print(f"{C.R}{t('error_network', e=e)}{C.W}")
        fresh = []

    new_count = 0
    for f in fresh:
        if f["ID"] not in by_id:
            by_id[f["ID"]] = f
            new_count += 1
        else:
            by_id[f["ID"]]["Title"] = f["Title"]   # cap nhat title moi nhat
            by_id[f["ID"]]["URL"] = f["URL"]
    data = list(by_id.values())

    logs, points = [], None
    if has_cookie:
        viewed = fetch_viewed_titles(session, has_cookie)
        matched = apply_viewed(data, viewed)
        try:
            logs = fetch_contribution_logs(session)
        except Exception:
            logs = []
        points = fetch_lifetime_points(session)
        if matched:
            print(f"{C.G}{t('sync_marked_viewed', n=matched)}{C.W}")

    save_excel(data)
    if new_count:
        print(f"{C.G}{t('sync_new_items', n=new_count)}{C.W}")
    return data, logs, points

# ----------------------------- Dashboard / de xuat -----------------------------
def today_log_stats(logs):
    """Tu log THAT cua hom nay: (videos, articles, daily_active, points)."""
    t = today_str()
    vids = arts = daily = pts = 0
    for l in logs:
        if (l.get("occurredAt") or "")[:10] != t:
            continue
        pts += l.get("point") or 0
        role = (l.get("contributionRole") or {}).get("name") or ""
        if role == "Watch a Video":
            vids += 1
        elif role == "Read Content":
            arts += 1
        elif role == "Daily Active":
            daily += 1
    return vids, arts, daily, pts

def dashboard(data, logs, points):
    videos   = [d for d in data if d["Type"] == "Video"]
    articles = [d for d in data if d["Type"] == "Article"]
    vv = sum(1 for d in videos   if d["Status"] == "Da xem")
    av = sum(1 for d in articles if d["Status"] == "Da xem")
    vt, at, daily, pts_today = today_log_stats(logs)

    print()
    print(f"{C.CY}{'='*48}{C.W}")
    print(f"{C.CY}{C.BOLD}        {t('dashboard_title')}{C.W}")
    print(f"{C.CY}{'='*48}{C.W}")

    print(f"\n{C.Y}{t('dashboard_lifetime_section')}{C.W}")
    lp = points if points is not None else "?"
    print(f"  {t('dashboard_lifetime_points'):<16}: {C.M}{lp}{C.W}")
    print(f"  {t('dashboard_videos_progress'):<16}: {vv:3d} / {len(videos):<3d} {t('dashboard_viewed_suffix')}")
    print(f"  {t('dashboard_articles_progress'):<16}: {av:3d} / {len(articles):<3d} {t('dashboard_read_suffix')}")

    print(f"\n{C.Y}{t('dashboard_today_section', date=today_str())}{C.W}")
    vc = C.G if vt >= VIDEO_DAILY_MAX else C.W
    ac = C.G if at >= ARTICLE_DAILY_MAX else C.W
    print(f"  {vc}{t('dashboard_video_quota_today')} {vt}/{VIDEO_DAILY_MAX}{C.W}  |  "
          f"{ac}{t('dashboard_article_quota_today')} {at}/{ARTICLE_DAILY_MAX}{C.W}  |  {t('dashboard_daily_active')} {daily}")
    print(f"  {C.M}{t('dashboard_points_today', pts=pts_today)}{C.W}")

    vslot = max(0, VIDEO_DAILY_MAX - vt)
    aslot = max(0, ARTICLE_DAILY_MAX - at)
    earnable = vslot * VIDEO_POINTS + aslot * ARTICLE_POINTS
    print(f"\n{C.Y}{t('dashboard_max_potential')}{C.W}")
    print(f"  {C.G}{t('dashboard_earnable', vslot=vslot, aslot=aslot, earnable=earnable)}{C.W}")

def is_blacklisted(d):
    return str(d.get("Blacklist", "")).strip().lower() == "yes"

def get_recommendations(data, logs):
    """Tinh danh sach de xuat hom nay. Tra ve (videos, articles, vslot, aslot).
    LOAI BO item da blacklist (link die da report) -> khong bao gio de xuat lai."""
    vt, at, _, _ = today_log_stats(logs)
    vslot = max(0, VIDEO_DAILY_MAX - vt)
    aslot = max(0, ARTICLE_DAILY_MAX - at)

    def pick(ctype, slot):
        if slot <= 0:
            return []
        return sorted(
            [d for d in data
             if d["Type"] == ctype and d["Status"] == "Chua xem" and not is_blacklisted(d)],
            key=lambda d: d.get("Title", ""))[:slot]

    return pick("Video", vslot), pick("Article", aslot), vslot, aslot

def recommend(data, logs):
    vids, arts, vslot, aslot = get_recommendations(data, logs)

    print(f"\n{C.Y}{t('recommend_header')}{C.W}")
    print(f"{C.GR}{'-'*48}{C.W}")
    if vslot == 0 and aslot == 0:
        print(f"{C.G}{t('recommend_maxed_out')}{C.W}")
        return

    if vslot > 0:
        print(f"\n{C.CY}{t('recommend_videos_left', vslot=vslot)}{C.W}")
        if not vids:
            print(f"{C.GR}{t('recommend_no_videos')}{C.W}")
        for i, d in enumerate(vids, 1):
            print(f"  {i}. {d['Title']}")
            print(f"     {C.B}{d['URL']}{C.W}")

    if aslot > 0:
        print(f"\n{C.CY}{t('recommend_articles_left', aslot=aslot)}{C.W}")
        if not arts:
            print(f"{C.GR}{t('recommend_no_articles')}{C.W}")
        for i, d in enumerate(arts, 1):
            print(f"  {i}. {d['Title']}")
            print(f"     {C.B}{d['URL']}{C.W}")

    print(f"\n{C.M}{t('recommend_earnable', pts=vslot*4 + aslot*2)}{C.W}")

def report_dead_links(data, logs):
    """Hien danh sach item dang de xuat hom nay (video + article, danh so chung).
    Cho nhap so (vd '1,3,5') de danh dau Blacklist=Yes -> khong de xuat lai nua."""
    vids, arts, _, _ = get_recommendations(data, logs)
    items = vids + arts   # danh so chung: video truoc, article sau

    print(f"\n{C.Y}{t('report_header')}{C.W}")
    print(f"{C.GR}{'-'*48}{C.W}")
    if not items:
        print(f"{C.GR}{t('report_no_items')}{C.W}")
        return

    for i, d in enumerate(items, 1):
        print(f"  {i}. [{d['Type']}] {d['Title']}")
        print(f"     {C.B}{d['URL']}{C.W}")

    raw = input(f"\n{t('report_select_prompt')}").strip()
    if not raw:
        print(f"{C.GR}{t('report_nothing')}{C.W}")
        return

    picked = set()
    for tok in raw.replace(" ", "").split(","):
        if tok.isdigit():
            n = int(tok)
            if 1 <= n <= len(items):
                picked.add(n)

    n_black = 0
    for n in picked:
        d = items[n - 1]
        if not is_blacklisted(d):
            d["Blacklist"] = "Yes"
            n_black += 1

    if n_black:
        save_excel(data)
        print(f"{C.G}{t('report_confirmed', n=n_black)}{C.W}")
    else:
        print(f"{C.GR}{t('report_none_new')}{C.W}")

# ----------------------------- Menu -----------------------------
def menu_text():
    return (f"\n{C.CY}============ {t('menu_title')} ============{C.W}\n"
            f"  1. {t('menu_recommend')}\n"
            f"  2. {t('menu_report')}\n"
            f"  3. {t('menu_dashboard')}\n"
            f"  0. {t('menu_exit')}\n"
            f"{C.CY}=============================={C.W}")

def main():
    select_language()   # lan dau hoi EN/VI + luu config/language.txt; lan sau dung luon
    print(f"{C.CY}{t('startup')}{C.W}")
    print(f"{C.GR}{t('excel_path', path=EXCEL_PATH)}{C.W}")
    session, has_cookie = make_session()

    if not ensure_nonce(session):
        print(f"{C.R}{t('error_nonce')}{C.W}")
        return
    if has_cookie:
        print(f"{C.G}{t('cookie_loaded')}{C.W}")
    else:
        print(f"{C.Y}{t('no_cookie_public')}{C.W}")

    data, logs, points = sync(session, has_cookie)
    dashboard(data, logs, points)
    recommend(data, logs)

    if not sys.stdin.isatty():
        return

    while True:
        print(menu_text())
        c = input(t("prompt_choose")).strip()
        if c == "1":
            data, logs, points = sync(session, has_cookie)
            dashboard(data, logs, points)
            recommend(data, logs)
        elif c == "2":
            report_dead_links(data, logs)
        elif c == "3":
            dashboard(data, logs, points)
        elif c == "0":
            print(f"{C.CY}{t('exit_message')}{C.W}")
            break
        else:
            print(f"{C.R}{t('prompt_invalid')}{C.W}")

if __name__ == "__main__":
    main()
